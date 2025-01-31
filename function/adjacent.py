import openpyxl
import numpy as np

def adjmatrix(adj_path):
    '''
    Goal:
    Reads an Excel file containing an adjacency matrix, ensures the matrix is square,
    and makes it symmetric if it is not already.
    ---------------------------------------
    Parameters:
    adj_path (str): The file path to the Excel file containing the adjacency matrix.
    ---------------------------------------
    Returns:
    numpy.ndarray: A symmetric adjacency matrix.
    ---------------------------------------
    Raises:
    ValueError: If the matrix is not square.
    ---------------------------------------
    Notes:
    - This function uses openpyxl to read the Excel file instead of pandas' read_excel 
      due to an unspecified issue with pd.read_excel().
    - The function first checks if the matrix is already symmetric. If it is not, it 
      makes the matrix symmetric by copying the upper triangle to the lower triangle 
      and vice versa.
    ---------------------------------------
    NOTICE: pd.read_excel() cannot work properly here for some unknown reason.
    '''
    workbook = openpyxl.load_workbook(adj_path)
    sheet = workbook['Sheet1']
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))
    matrix = np.array(data)
    
    if matrix.shape[0] != matrix.shape[1]:
        raise ValueError("The matrix is not square")
    
    if np.array_equal(matrix, matrix.T):
        print("The matrix is already symmetric.")
        return matrix
    
    # upper to lower
    for i in range(matrix.shape[0]):
        for j in range(i + 1, matrix.shape[1]):
            matrix[j, i] = matrix[i, j]
    # insurance, lower to upper
    for i in range(matrix.shape[0]):
        for j in range(i):
            matrix[i, j] = matrix[j, i]
    
    return matrix

if __name__ == '__main__':
    workbook = openpyxl.load_workbook('D:\\STUDY\\CFPS\\geo\\adjacent\\adjacent.xlsx')
    sheet = workbook['Sheet1']
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))
    matrix = np.array(data)
    print(matrix.shape)
    
    print(adjmatrix('D:\\STUDY\\CFPS\\geo\\adjacent\\adjacent.xlsx').shape)
        